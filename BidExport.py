import tkinter as tk
from tkinter import simpledialog, messagebox
from selenium import webdriver
from selenium.webdriver.edge.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
from urllib.parse import urlparse
from datetime import datetime
import pandas as pd
from bs4 import BeautifulSoup
import re
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font

def parse_successful_bids(bid_list):
    """从竞价列表中提取成功竞价的信息"""
    successful_bids = []
    for li in bid_list.find_elements(By.TAG_NAME, "li"):
        try:
            if "竞价成功" in li.find_element(By.CLASS_NAME, "jl_1").text:
                bid_status = li.find_element(By.CLASS_NAME, "jl_1").text.strip()  # 竞价状态
                bidder = li.find_elements(By.CLASS_NAME, "jl_2")[0].text.strip()  # 出价人
                bid_time = li.find_elements(By.CLASS_NAME, "jl_2")[1].text.strip()  # 出价时间
                bid_amount = li.find_element(By.CLASS_NAME, "jl_3").text.strip()  # 出价金额
                successful_bids.append({"出价状态": bid_status, "出价人": bidder, "出价时间": bid_time, "出价金额": bid_amount})
        except Exception:
            continue
    return successful_bids

def parse_unsuccessful_bids(bid_list, successful_bids_data):
    """从竞价列表中提取未成功竞价的信息"""
    unsuccessful_bids = []
    for li in bid_list.find_elements(By.TAG_NAME, "li"):
        try:
            bidder = li.find_elements(By.CLASS_NAME, "jl_2")[0].text.strip()    # 出价人
            bid_status = li.find_element(By.CLASS_NAME, "jl_1").text.strip()    # 竞价状态
            # if "竞价成功" not in bid_status and check_bid_exist(successful_bids_data, bidder) and check_bid_exist(unsuccessful_bids_data, bidder):
            if "竞价成功" not in bid_status and check_bid_exist(successful_bids_data, bidder):
                bid_time = li.find_elements(By.CLASS_NAME, "jl_2")[1].text.strip()  # 出价时间
                bid_amount = li.find_element(By.CLASS_NAME, "jl_3").text.strip()  # 出价金额
                unsuccessful_bids.append({"出价状态": "竞价失败", "出价人": bidder, "出价时间": bid_time, "出价金额": bid_amount})
        except Exception:
            continue
    return unsuccessful_bids

def check_bid_exist(bids_data, bidder):
    # print(bids_data)
    if len(bids_data) == 0:
        return True
    for bid in bids_data:
        if bidder in bid['出价人']:
            # 如果出价人已经存在，则返回False
            return False
    # 如果出价人不存在，则返回True
    return True

def deduplication(bids_data):
    unsuccessful_bids = []
    
    for bid in bids_data:
        bidder = bid['出价人']
        bid_amount = bid['出价金额']
        
        # 检查该出价人是否已存在于未成功竞价列表中
        existing_bid = next((item for item in unsuccessful_bids if item['出价人'] == bidder), None)
        
        if existing_bid:
            # 如果该出价人的出价金额更高，更新记录
            if bid_amount > existing_bid['出价金额']:
                existing_bid.update(bid)
        else:
            # 如果该出价人不存在，直接添加新记录
            unsuccessful_bids.append(bid)
    
    return unsuccessful_bids


def get_item_name(driver):
    """获取商品名称"""
    item_name_element = driver.find_element(By.CLASS_NAME, "i_tit")
    return item_name_element.text.strip()

def auto_bid_until_end(driver, target_successful_count, bid_type, theater_name):
    successful_bids_data = []  # 用于存储所有竞价成功信息
    unsuccessful_bids_data = []  # 用于存储所有竞价未成功信息
    total_bids_data = []
    wait = WebDriverWait(driver, 10)

    # 获取商品名称
    # item_name = get_item_name(driver)
    
    max_page_element = driver.find_element(By.XPATH, '//*[@id="d_blist"]/div[4]/span[3]')
    max_page = int(max_page_element.text)
    now_page = 1
    
    #保存竞价成功的数据
    while now_page <= max_page:
        # 获取当前页的成功竞价信息
        u_blist = driver.find_element(By.ID, "u_blist")
        successful_bids_data.extend(parse_successful_bids(u_blist))
        
        if len(successful_bids_data) >= target_successful_count:
            print("已获取所有成功竞价信息")
            break
        
        # 检查 u_blistM 列表是否有额外成功竞价信息
        u_blistM = driver.find_element(By.ID, "u_blistM")
        successful_bids_data.extend(parse_successful_bids(u_blistM))
        
        if len(successful_bids_data) >= target_successful_count:
            print("已获取所有成功竞价信息")
            break
        
        # 翻页
        next_button = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="a_b_n"]')))
        next_button.click()
        print(f"加载第 {now_page} 页...")
        now_page += 1
        time.sleep(0.15)
    
    #保存竞价未成功的数据
    while now_page <= max_page:
        # 获取当前页的成功竞价信息
        u_blist = driver.find_element(By.ID, "u_blist")
        unsuccessful_bids_data.extend(parse_unsuccessful_bids(u_blist, successful_bids_data))
        
        # 检查 u_blistM 列表是否有额外成功竞价信息
        u_blistM = driver.find_element(By.ID, "u_blistM")
        unsuccessful_bids_data.extend(parse_unsuccessful_bids(u_blistM, successful_bids_data))
        
        # 翻页
        next_button = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="a_b_n"]')))
        next_button.click()
        print(f"加载第 {now_page} 页...")
        now_page += 1
        time.sleep(0.15)
    
    unsuccessful_bids_data = deduplication(unsuccessful_bids_data)
    total_bids_data = successful_bids_data + unsuccessful_bids_data

    #获取座位号的数组
    seats = get_seat_positon(theater_name, bid_type, target_successful_count)
    # 为每条竞价记录按剧场分配座位号
    for idx, bid in enumerate(total_bids_data):
        bid["座位类型"] = bid_type
        #判断剧场
        # if "SNH" in theater_name:
        #     seat_number = get_seat_positon_SNH(bid_type, idx)
        # elif "HGH" in theater_name:
        #     seat_number = get_seat_positon_HGH(bid_type, idx)
        # elif "BEJ" in theater_name:
        #     seat_number = get_seat_positon_BEJ(bid_type, idx)
        # elif "MINILIVE" in theater_name:
        #     seat_number = get_seat_positon_MiniLive(idx)
        # elif "拍立得" in theater_name:
        #     seat_number = get_seat_positon_pld(target_successful_count, idx)
        if(idx > len(seats) - 1):
            bid["座位号"] = "竞价失败"
        else:
            bid["座位号"] = seats[idx]  # 添加座位号

    return total_bids_data

# 定义一个函数来计算站区座位数
def calculate_seat_count(pattern_1, pattern_2, text):
    # 匹配单一区间
    matches_1 = re.findall(pattern_1, text)
    # 匹配多个区间
    matches_2 = re.findall(pattern_2, text)

    total_seats = 0
    print(matches_1, matches_2)
    # 如果有单一区间
    if matches_1:
        for match in matches_1:
            start, end = map(int, match)  # 转换为整数
            total_seats += (end - start + 1)  # 计算区间内的座位数

    # 如果有多个区间
    elif matches_2:
        for match in matches_2:
            start, end = map(int, match[:2])  # 第一个区间
            total_seats += (end - start + 1)  # 计算第一个区间的座位数

            # 处理第二个区间（如果存在）
            if match[2] and match[3]:
                start2, end2 = map(int, match[2:4])  # 第二个区间
                total_seats += (end2 - start2 + 1)

    return total_seats


#获取SNH剧场竞价门票数量
def get_bid_number_SNH(bid_type, driver):
    item_info = driver.find_element(By.XPATH, '//*[@id="TabTab03Con1"]').get_attribute('outerHTML')
    soup = BeautifulSoup(item_info, 'html.parser')

    # 获取整个文本内容
    full_text = soup.get_text()

    if "普站" in bid_type:
        # 正则表达式匹配座位区间（例如“025至100”或“025至30、36至100”）
        seat_pattern_1 = r"站区序号(\d{3})至(\d{3})"  # 单一区间（如：025至100）
        seat_pattern_2 = r"站区序号(\d{3})至(\d{2})(?:、(\d{2})至(\d{3}))*"  # 多个区间（如：025至30、36至100）
        seat_count = calculate_seat_count(seat_pattern_1, seat_pattern_2, full_text)
        return seat_count


    #如果是生公
    if "生日潮流包" in full_text:
        # 根据商品名称判断座位类型
        if "SVIP" in bid_type:
            return 24
        elif "VIP" in bid_type:
            return 79
        elif "摄影" in bid_type:
            return 24
        elif "杆位" in bid_type:
            return 24
        elif "超级" in bid_type:
            return 3
        elif "普座" in bid_type:
            return 122
        return 0
    else:
        # 根据商品名称判断座位类型
        if "SVIP" in bid_type:
            return 24
        elif "VIP" in bid_type:
            return 84
        elif "摄影" in bid_type:
            return 24
        elif "杆位" in bid_type:
            return 24
        elif "超级" in bid_type:
            return 3
        elif "普座" in bid_type:
            return 132
        return 0



#获取杭州剧场竞价门票数量
def get_bid_number_HGH(bid_type):
    # 根据商品名称判断座位类型
    if "超级" in bid_type:
        return 54
    elif "摄影" in bid_type:
        return 19
    return 0

#获取BEJ剧场竞价门票数量
def get_bid_number_BEJ(driver):
    item_info = driver.find_element(By.XPATH, '//*[@id="TabTab03Con1"]').get_attribute('outerHTML')
    soup = BeautifulSoup(item_info, 'html.parser')

    # 获取整个文本内容
    full_text = soup.get_text()

    # 使用正则表达式提取所有票数信息，只需包含 "演出门票"
    ticket_counts = re.findall(r".*?演出门票.*?(\d+)张", full_text)

    # 如果找到了票数信息
    if ticket_counts:
        total_tickets = int(ticket_counts[0])  # 取第一个匹配的票数
        return total_tickets
    else:
        return 0
    
#获取MINILIVE竞价门票数量
def get_bid_number_MiniLive(driver):
    span_element = driver.find_element(By.XPATH, '//*[@id="TabTab03Con1"]/span')

    # 使用正则表达式提取所有票数信息，只需包含 "演出门票"
    ticket_counts = re.findall(r".*?入场资格(\d+)位", span_element.text)

    # 如果找到了票数信息
    if ticket_counts:
        total_tickets = int(ticket_counts[0])  # 取第一个匹配的票数
        return total_tickets
    else:
        return 0
    
#获取竞价拍立得数量
def get_bid_number_pld(driver):
    item_info = driver.find_element(By.XPATH, '//*[@id="TabTab03Con1"]').get_attribute('outerHTML')
    soup = BeautifulSoup(item_info, 'html.parser')

    # 获取整个文本内容
    full_text = soup.get_text()

    # 使用正则表达式提取所有票数信息，只需包含 "演出门票"
    ticket_counts = re.findall(r".*?共.*?(\d+)套", full_text)

    # 如果找到了票数信息
    if ticket_counts:
        total_tickets = int(ticket_counts[0])  # 取第一个匹配的票数
        return total_tickets
    else:
        return 0
    
    #获取竞价拍立得数量
def get_bid_number_birthparty(driver, theater_name):
    item_info = driver.find_element(By.XPATH, '//*[@id="TabTab03Con1"]').get_attribute('outerHTML')
    soup = BeautifulSoup(item_info, 'html.parser')

    # 获取整个文本内容
    full_text = soup.get_text()
    if("SNH" in theater_name):
        # 使用正则表达式提取所有票数信息，只需包含 "演出门票"
        ticket_counts = re.findall(r".*?名额：.*?(\d+)名", full_text)
        # 如果找到了票数信息
        if ticket_counts:
            total_tickets = int(ticket_counts[0])  # 取第一个匹配的票数
            return total_tickets
        else:
            return 0
    elif("BEJ" in theater_name):
        # 使用正则表达式提取所有票数信息，只需包含 "演出门票"
        ticket_counts = re.findall(r".*?竞拍数量：.*?(\d+)张", full_text)
        # 如果找到了票数信息
        if ticket_counts:
            total_tickets = int(ticket_counts[0])  # 取第一个匹配的票数
            return total_tickets
        else:
            return 0


def get_seat_type(item_name):
    """根据商品名称判断座位类型"""
    # 根据商品名称判断座位类型
    if "超级" in item_name:
        return "超级"
    elif "SVIP" in item_name:
        return "SVIP"
    elif "摄影" in item_name:
        return "摄影"
    elif "杆位" in item_name:
        return "杆位"
    elif "普站" in item_name:
        return "普站"
    elif "VIP" in item_name:
        return "VIP"
    elif "普座" in item_name:
        return "普座"  
    elif "MINILIVE" in item_name:
        return "MINILIVE"
    elif "拍立得" in item_name:
        return "拍立得"
    elif "生日会" in item_name:
        return "生日会"
    else:
        return "其他"


def get_seat_positon(theater_name, bid_type, bid_count = 0):
    if "SNHbirthday" in theater_name and bid_count == 71 and bid_type == "普站":
        return get_seat_positon_SNH_birthday(bid_type)
    if "SNHbirthday" in theater_name and bid_count == 76 and bid_type == "普站":
        return get_seat_positon_SNH(bid_type)
    elif "SNHbirthday" in theater_name:
        return get_seat_positon_SNH_birthday(bid_type)
    elif "SNH" in theater_name :
        return get_seat_positon_SNH(bid_type)
    elif "HGH" in theater_name:
        return get_seat_positon_HGH(bid_type)
    elif "BEJ" in theater_name:
        return get_seat_positon_BEJ(bid_type)
    elif "MINILIVE" in theater_name:
        return get_seat_positon_MiniLive(bid_count)
    elif "拍立得" in theater_name:
        return get_seat_positon_pld(bid_count)
    elif "生日会" in theater_name:
        return get_seat_positon_birthparty(bid_count)


def get_seat_positon_SNH(bid_type):
    """根据竞价类型和索引为每个竞价分配座位号"""
    seats = []
    
    # 普座
    if bid_type == "普座":
        rows_6_col_1_18 = [f"6排{j}" for j in range(1, 19)]  # 6排1到6排18
        rows_5_6_col_19_20 = [f"{i}排{j}" for i in range(5, 7) for j in range(19, 21)]  # 5排18、19 | 6排18、19
        rows_4_6_col_21_22 = [f"{i}排{j}" for i in range(4, 7) for j in range(21, 23)]  # 4排21、22 | 5排21、22 | 6排21、22
        rows_3_6_col_23_24 = [f"{i}排{j}" for i in range(3, 7) for j in range(23, 25)]  # 3排23、24 | 4排23、24 | 5排23、24 | 6排23、24
        rows_7_10 = [f"{i}排{j}" for i in range(7, 11) for j in range(1, 25)]  # 7排到10排
        seats = rows_6_col_1_18 + rows_5_6_col_19_20 + rows_4_6_col_21_22 + rows_3_6_col_23_24 + rows_7_10  # 普座座位
    elif bid_type == "SVIP":
        seats = [f"1排{i}" for i in range(1, 25)]  # 摄影座位
    elif bid_type == "VIP":
        seats = [f"{i}排{j}" for i in range(2, 6) for j in range(1, 25 - 2 * (i - 2))]  # VIP座位
        # seats = seats[:84]  # 限制为84个
    elif bid_type == "摄影":
        seats = [f"1排{i}" for i in range(1, 25)]  # 摄影座位
    elif bid_type == "杆位":
        seats = [str(i) for i in range(1, 25)]  # 杆位座位
    elif bid_type == "普站":
        seats = [str(i) for i in range(25, 101)]  # 普站座位
    elif bid_type == "超级":
        seats = ["中", "左", "右"] 
    
    return seats

def get_seat_positon_SNH_birthday(bid_type):
    """根据竞价类型和索引为每个竞价分配座位号"""
    seats = []
    # 普座
    if bid_type == "普座":
        rows_6_col_1_18 = [f"6排{j}" for j in range(1, 19)]  # 6排1到6排18
        rows_5_6_col_19_20 = [f"{i}排{j}" for i in range(5, 7) for j in range(19, 21)]  # 5排18、19 | 6排18、19
        rows_4_6_col_21_22 = [f"{i}排{j}" for i in range(4, 7) for j in range(21, 23)]  # 4排21、22 | 5排21、22 | 6排21、22
        rows_3_6_col_23_24 = [f"{i}排{j}" for i in range(3, 7) for j in range(23, 25)]  # 3排23、24 | 4排23、24 | 5排23、24 | 6排23、24
        rows_7_col_1_19 = [f"7排{j}" for j in range(1, 21) if j % 2 != 0]   #7排1、3、5、7、9、11、13、15、17、19
        rows_7_col_21_24 = [f"7排{j}" for j in range(21, 25)]           #7排21、22、23、24
        rows_8_10 = [f"{i}排{j}" for i in range(8, 11) for j in range(1, 25)]  # 8排到10排
        seats = rows_6_col_1_18 + rows_5_6_col_19_20 + rows_4_6_col_21_22 + rows_3_6_col_23_24 + rows_7_col_1_19 + rows_7_col_21_24 + rows_8_10  # 普座座位
    elif bid_type == "VIP":
        rows_2_col_2_10 = [f"2排{j}" for j in range(1, 11) if j % 2 == 0]
        rows_2_col_11_24 = [f"2排{j}" for j in range(11, 25)]
        rows_3_5 = [f"{i}排{j}" for i in range(3, 6) for j in range(1, 25 - 2 * (i - 2))]  # VIP座位
        seats = rows_2_col_2_10 + rows_2_col_11_24 + rows_3_5
        # seats = seats[:84]  # 限制为84个
    elif bid_type == "摄影":
        seats = [f"1排{i}" for i in range(1, 25)]  # 摄影座位
    elif bid_type == "杆位":
        seats = [str(i) for i in range(1, 25)]  # 杆位座位
    elif bid_type == "普站":
        stand_25_30 = [str(i) for i in range(25, 31)]  # 普站座位
        stand_31_100 = [str(i) for i in range(36, 101)]
        seats = stand_25_30 + stand_31_100
    elif bid_type == "超级":
        seats = ["中", "左", "右"]
    
    return seats

def get_seat_positon_HGH(bid_type):
    """根据竞价类型和索引为每个竞价分配座位号"""
    seats = []
    
    # 普座
    if bid_type == "超级":
        rows_1_col_1_25 = [f"1排{j}" for j in range(1, 26)]  # 1排1到1排25
        rows_2_col_1_29 = [f"2排{j}" for j in range(1, 30)]  # 2排1到2排29
        seats = rows_1_col_1_25 + rows_2_col_1_29  # 普座座位
    elif bid_type == "摄影":
        seats = [f"11排{i}" for i in range(1, 20)]  # 摄影座位
    
    return seats

def get_seat_positon_BEJ(bid_type):
    """根据竞价类型和索引为每个竞价分配座位号"""
    seats = []
    
    # 普座
    if bid_type == "超级":
        seats = ["中", "左", "右"]
    elif bid_type == "VIP":
        rows_1_4 = [f"{i}排{j}" for i in range(1, 5) for j in range(1, 18)]  # VIP座位
        rows_5 = [f"5排3"] + [f"5排{j}" for j in range(5, 18)]  # VIP座位
        rows_6 = [f"6排13"] + [f"6排{j}" for j in range(15, 18)]  # VIP座位
        seats = rows_1_4 + rows_5 + rows_6  # 限制为84个
    elif bid_type == "摄影":
        # rows_6_col_5_12 = [f"6排{i}" for i in range(5, 13)]  # 摄影座位
        seats = [f"6排3"]+ [f"6排6"] + [f"6排5"] + [f"6排8"] + [f"6排7"] + [f"6排10"] + [f"6排9"] + [f"6排12"] + [f"6排11"] + [f"6排14"]
    
    return seats

def get_seat_positon_MiniLive(bid_number):
    """根据竞价类型和索引为每个竞价分配座位号"""
    seats = []
    seats = [str(i) for i in range(1, bid_number + 1)]  # MINILIVE座位
    
    return seats

def get_seat_positon_pld(bid_number):
    """根据竞价类型和索引为每个竞价分配座位号"""
    seats = []
    seats = [str(i) for i in range(1, bid_number + 1)]  # 拍立得位置
    
    return seats

def get_seat_positon_birthparty(bid_number):
    """根据竞价类型和索引为每个竞价分配座位号"""
    seats = []
    seats = [str(i) for i in range(1, bid_number + 1)]  # 冷餐座位
    
    return seats

def save_excel(successful_bids_data, item_name, output_file="bidding_results.xlsx"):

    # 保存数据到 Excel 文件
    df = pd.DataFrame(successful_bids_data)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Bidding Results"
    
    # 写入商品名称
    ws.append([item_name])
    
    # 写入表头并加粗
    header = ["出价状态", "出价人", "出价时间", "出价金额", "座位类型", "座位号"]
    ws.append(header)
    for cell in ws[2]:  # 第二行是标题
        cell.font = Font(bold=True)
    
    # 写入竞价数据
    for row in dataframe_to_rows(df, index=False, header=False):
        ws.append(row)
    
    #更新最小、最大出价、最早、最晚出价的出价人信息
    ws = update_min_max_info(df, ws)

    wb.save(output_file + ".xlsx")
    print(f"竞价成功信息已保存至 {output_file}" + ".xlsx")
    

def update_min_max_info(df, ws):

    # 只选择出价状态为 "竞价成功" 的记录
    df_successful = df[df['出价状态'] == '竞价成功']

    # 将 '出价时间' 列转换为 datetime 类型，处理格式问题
    # df_successful['出价时间'] = pd.to_datetime(df_successful['出价时间'], format='%Y/%m/%d %H:%M:%S', errors='coerce')
    # df_successful['出价金额'] = pd.to_numeric(df_successful['出价金额'], errors='coerce')
    df_successful.loc[:, '出价时间'] = pd.to_datetime(df_successful['出价时间'], format='%Y/%m/%d %H:%M:%S', errors='coerce')
    df_successful.loc[:, '出价金额'] = pd.to_numeric(df_successful['出价金额'], errors='coerce')

    # 提取最早出价时间、最晚出价时间、最高出价、最低出价
    earliest_bid_time = df_successful['出价时间'].min()
    latest_bid_time = df_successful['出价时间'].max()
    highest_bid = df_successful['出价金额'].max()
    lowest_bid = df_successful['出价金额'].min()

    # 获取相应的人的详细信息
    earliest_bidder = df_successful[df_successful['出价时间'] == earliest_bid_time].iloc[0]
    latest_bidder = df_successful[df_successful['出价时间'] == latest_bid_time].iloc[0]
    highest_bidder = df_successful[df_successful['出价金额'] == highest_bid].iloc[0]
    lowest_bidder = df_successful[df_successful['出价金额'] == lowest_bid].iloc[0]

    # 将这些值插入到Excel的最后四行
    row_index = len(df) + 3  # 在最后四行之前插入

    # 设置加粗
    bold_font = Font(bold=True)

    ws[f'A{row_index + 1}'].font = bold_font
    ws[f'A{row_index + 2}'].font = bold_font
    ws[f'A{row_index + 3}'].font = bold_font
    ws[f'A{row_index + 4}'].font = bold_font

    # 插入每个人的详细数据，最早出价的那个人的详细数据
    row_number = row_index + 1
    ws[f'A{row_number}'] = '最早出价者'
    ws[f'B{row_number}'] = earliest_bidder['出价人']
    ws[f'C{row_number}'] = earliest_bidder['出价时间'].strftime('%Y-%m-%d %H:%M:%S')
    ws[f'D{row_number}'] = earliest_bidder['出价金额']
    ws[f'E{row_number}'] = earliest_bidder['座位类型']
    ws[f'F{row_number}'] = earliest_bidder['座位号']

    # 插入最晚出价的那个人的详细数据
    row_number += 1
    ws[f'A{row_number}'] = '最晚出价者'
    ws[f'B{row_number}'] = latest_bidder['出价人']
    ws[f'C{row_number}'] = latest_bidder['出价时间'].strftime('%Y-%m-%d %H:%M:%S')
    ws[f'D{row_number}'] = latest_bidder['出价金额']
    ws[f'E{row_number}'] = latest_bidder['座位类型']
    ws[f'F{row_number}'] = latest_bidder['座位号']

    # 插入最高出价的那个人的详细数据
    row_number += 1
    ws[f'A{row_number}'] = '最高出价者'
    ws[f'B{row_number}'] = highest_bidder['出价人']
    ws[f'C{row_number}'] = highest_bidder['出价时间'].strftime('%Y-%m-%d %H:%M:%S')
    ws[f'D{row_number}'] = highest_bidder['出价金额']
    ws[f'E{row_number}'] = highest_bidder['座位类型']
    ws[f'F{row_number}'] = highest_bidder['座位号']

    # 插入最低出价的那个人的详细数据
    row_number += 1
    ws[f'A{row_number}'] = '最低出价者'
    ws[f'B{row_number}'] = lowest_bidder['出价人']
    ws[f'C{row_number}'] = lowest_bidder['出价时间'].strftime('%Y-%m-%d %H:%M:%S')
    ws[f'D{row_number}'] = lowest_bidder['出价金额']
    ws[f'E{row_number}'] = lowest_bidder['座位类型']
    ws[f'F{row_number}'] = lowest_bidder['座位号']

    return ws


def get_item_id(url):
    path = urlparse(url).path
    # 获取路径中的最后一个部分作为 item_id
    item_id = path.split('/')[-1]
    return item_id

def stats_one_good(driver):
    theater_name = driver.find_element(By.XPATH, "/html/body/div[2]/div/div[2]/div[2]/ul/li[2]/p").text
    excel_name = driver.find_element(By.XPATH, "/html/body/div[2]/div/div[2]/div[2]/ul/li[1]").text

    # 获取商品名称元素
    title_name_element = driver.find_element(By.CLASS_NAME, "i_tit")
    title_name = title_name_element.text.strip()  # 获取文本内容

    #获取商品详细信息并判断是否为生公
    item_info = driver.find_element(By.XPATH, '//*[@id="TabTab03Con1"]').get_attribute('outerHTML')
    soup = BeautifulSoup(item_info, 'html.parser')

    # 获取整个文本内容
    item_info_text = soup.get_text()
    if("生日潮流包" in item_info_text):
        birthday = True
    else:
        birthday = False


    bid_type = get_seat_type(title_name)
    if("SNH" in theater_name and "星梦剧院" in title_name and "MINILIVE" not in title_name):
        bid_number = get_bid_number_SNH(bid_type, driver)
        if(birthday):
            theater_name = "SNHbirthday"
    elif("SNH" in theater_name and "星梦空间" in title_name and "MINILIVE" not in title_name):
        bid_number = get_bid_number_HGH(bid_type)
        theater_name = "HGH"
    elif("BEJ" in theater_name and "生日会" not in title_name):
        bid_number = get_bid_number_BEJ(driver)
    elif("MINILIVE" in title_name):
        bid_number = get_bid_number_MiniLive(driver)
        theater_name = "MINILIVE"
    elif("拍立得" in title_name):
        bid_number = get_bid_number_pld(driver)
        theater_name = "拍立得"
    elif("生日会" in title_name):
        bid_number = get_bid_number_birthparty(driver, theater_name)
        theater_name = "生日会"
    
    print(f"一共有{bid_number}个位置")

    if(bid_number != 0):
        max_bid_num = bid_number

    total_bids_data = []  # 用于存储所有竞价成功信息
    total_bids_data = auto_bid_until_end(driver, max_bid_num, bid_type, theater_name)

    # 获取商品名称
    item_id = get_item_name(driver)
    save_excel(total_bids_data, item_id, excel_name)
    driver.refresh()


# 创建 GUI 窗口
root = tk.Tk()
root.withdraw()  # 隐藏主窗口

# 获取屏幕宽度和高度
screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()

# 设置窗口大小
window_width = 400
window_height = 300

# 计算居中位置
x = (screen_width // 2) - (window_width // 2)
y = (screen_height // 2) - (window_height // 2)

# 设置窗口位置和大小
root.geometry(f"{window_width}x{window_height}+{x}+{y}")

# 弹出输入框获取用户输入
# edge_driver_path = simpledialog.askstring("输入", "请输入 Edge WebDriver 路径：", initialvalue="D:\PythonProgram\msedgedriver.exe")
target_url = simpledialog.askstring("输入", "请输入目标网站 URL：")
# username = simpledialog.askstring("输入", "请输入账户：", initialvalue="13916294173")
# password = simpledialog.askstring("输入", "请输入密码：", show='*', initialvalue="Guan0527")  # 密码框，输入的字符将显示为星号

edge_driver_path = "D:\PythonProgram\msedgedriver.exe"
username = "13916294173"
password = "Guan0527"  

#输入最大竞价人数
# max_bid_num = simpledialog.askinteger("输入", "请输入最大竞价人数（默认为46）：", initialvalue=46)
max_bid_num = 46


# 关闭 GUI 窗口
root.quit()

# 确保输入不为空
if not all([edge_driver_path, target_url, username, password, max_bid_num]):
    messagebox.showerror("错误", "所有输入框都必须填写！")

# 设置Edge WebDriver路径
# edge_driver_path = 'D:\PythonProgram\msedgedriver.exe' # 替换为msedgedriver.exe的实际路径

# 创建 Edge 的 Service 对象
service = Service(executable_path=edge_driver_path)

# 初始化 Edge 浏览器
driver = webdriver.Edge(service=service)

try:
    
    # 打开目标网站
    driver.get(target_url)

    # 等待页面加载
    wait = WebDriverWait(driver, 1)  # 最多等待1秒

    # 登录
    try:
        login = wait.until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[1]/div/div[1]/div[1]/a[1]')))
        login.click()
    except:
        print("登录按钮不存在")

    # 选择账号密码登录
    login = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="login"]/div[1]/div/div[1]/div/ul/li[1]/a')))
    login.click()
        
    # 等待并输入账号
    account_input = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="username"]')))
    account_input.send_keys(username)

    # 输入密码
    password_input = driver.find_element(By.XPATH, '//*[@id="password"]')
    password_input.send_keys(password)

    # 点击登录按钮
    login_button = driver.find_element(By.XPATH, '//*[@id="loginbtn2"]')
    login_button.click()
    print("登录成功")

    # 等待页面返回目标 URL
    while True:
        current_url = driver.current_url
        if current_url.startswith(target_url):
            break
        time.sleep(1)  # 每秒检查一次，直到返回目标 URL

    # 获取商品id
    item_id = get_item_id(target_url)

    print("开始获取数据")
    time.sleep(1)

    #测试
    # item_info = driver.find_element(By.XPATH, '//*[@id="TabTab03Con1"]').get_attribute('outerHTML')
    # soup = BeautifulSoup(item_info, 'html.parser')

    # # 获取整个文本内容
    # full_text = soup.get_text()

    # # 使用正则表达式提取所有票数信息，只需包含 "演出门票"
    # ticket_counts = re.findall(r".*?演出门票.*?(\d+)张", full_text)

    # # 如果找到了票数信息
    # if ticket_counts:
    #     total_tickets = int(ticket_counts[0])  # 取第一个匹配的票数
    #     print(f"一共有 {total_tickets} 张票")
    # else:
    #     print("没有找到票数信息")
    # print(get_bid_number_pld(driver))
    #测试
    multiPage = True
    if("pageNum" not in target_url):
        multiPage = False
        total_pages = 1
    else:
        # 获取分页部分的HTML
        pagination_html = driver.find_element(By.ID, "pagination").get_attribute("innerHTML")

        # 使用BeautifulSoup解析HTML
        soup = BeautifulSoup(pagination_html, "html.parser")

        # 获取分页链接中最大的页码，这通常是总页数
        page_links = soup.find_all("a")
        # 过滤出所有有效的数字页码（排除 '>' 和 '<' 等非数字链接）
        page_numbers = []
        for link in page_links:
            try:
                # 提取页码并将其转为整数
                page_number = int(link.text)
                page_numbers.append(page_number)
            except ValueError:
                # 跳过非数字的链接（例如 '<', '>' 等）
                continue

        # 获取最大页码即总页数
        total_pages = max(page_numbers)

        #分割url以便获取每页url
        base_url_front = target_url.split("pageNum=")[0]
        base_url_end = target_url.split("pageNum=")[1].split("&")[1:]

        print(f"一共有 {total_pages} 页")
    #遍历每一页lian
    for page in range(0, total_pages):
        if(multiPage):
            page_url = f"{base_url_front}pageNum={page}&{ '&'.join(base_url_end)}"
            driver.get(page_url)
            # 等待页面加载
            wait = WebDriverWait(driver, 1)

        # 要访问的 URL数组
        good_urls = []

        # 获取所有商品的链接
        goods_links = driver.find_elements(By.CSS_SELECTOR, "div.goods a")
        for link in goods_links:
            # 获取商品页面的 URL
            good_url = link.get_attribute("href")
            good_urls.append(good_url)

        #去重
        good_urls = list(set(good_urls))
        print(f"urls：{good_urls}")

        if(good_urls != []):
            for good_url in good_urls:
                #进入具体竞价商品页面
                driver.get(good_url)

                # 等待页面加载
                wait = WebDriverWait(driver, 1)  # 最多等待1秒

                stats_one_good(driver)
        else:
            stats_one_good(driver)
        
        # next_page_button = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="pagination"]/div/a[3]')))
        # next_page_button.click()
    # time.sleep(5)  # 休眠几秒以查看结果

finally:
    # 关闭浏览器
    driver.quit()
    # time.sleep(5)

